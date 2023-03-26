#!/usr/bin/env python3
# coding: utf-8
import importlib
import sys
import yaml
import rospy

importlib.reload(sys)
import argparse
import csv
import math
import time
from collections import namedtuple
from math import pi
from os import path

import geometry_msgs.msg
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import plotly.graph_objs as go
import plotly.offline as py
import roboticstoolbox as rtb
import sympy as sp
# import dyna_space
from interface_control.msg import (cal_cmd, cal_process, cal_result, dyna_data,
                                dyna_space_data, optimal_design,
                                optimal_random, specified_parameter_design, tested_model_name, arm_structure)
from matplotlib import cm
from moveit_msgs.msg import DisplayTrajectory, RobotTrajectory
from mpl_toolkits.mplot3d import Axes3D
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, colors
from openpyxl.utils import get_column_letter
from plotly.offline import download_plotlyjs, iplot, plot
from scipy.interpolate import make_interp_spline  # draw smooth
from spatialmath import *
from trajectory_msgs.msg import JointTrajectory, JointTrajectoryPoint
# from dqn import DQN

np.set_printoptions(
    linewidth=100,
    formatter={"float": lambda x: f"{x:8.4g}" if abs(x) > 1e-10 else f"{0:8.4g}"},
)

from arm_workspace import arm_workspace_plane
# from robot_urdf import RandomRobot
from motor_module import motor_data
from random_robot import RandomRobot
from modular_robot_6dof import modular_robot_6dof
# DRL_optimization api
import sys
import os
import torch.nn as nn
import torch.nn.functional as F
curr_path = os.path.dirname(os.path.abspath(__file__))  # 当前文件所在绝对路径
parent_path = os.path.dirname(curr_path)  # 父路径
sys.path.append(parent_path)  # 添加路径到系统路径

import gym
import torch
import datetime
import numpy as np
from common.utils import save_results, make_dir
from common.utils import plot_rewards
from dqn import DQN

import argparse
import random

import matplotlib.pyplot as plt
# from RobotOptEnv_dynamixel import RobotOptEnv, RobotOptEnv_3dof, RobotOptEnv_5dof
from RobotOptEnv_dynamixel_v2 import RobotOptEnv, RobotOptEnv_3dof, RobotOptEnv_5dof
import tensorboardX
import yaml

# add
import tensorflow as tf
from tf_agents.agents.dqn.dqn_agent import DqnAgent, DdqnAgent
from tf_agents.networks.q_network import QNetwork
from tf_agents.agents.categorical_dqn import categorical_dqn_agent
from tf_agents.drivers import dynamic_step_driver
from tf_agents.environments import suite_gym
from tf_agents.environments import tf_py_environment
from tf_agents.eval import metric_utils
from tf_agents.metrics import tf_metrics
from tf_agents.networks import categorical_q_network
from tf_agents.policies import random_tf_policy
from tf_agents.replay_buffers import tf_uniform_replay_buffer
from tf_agents.trajectories import trajectory
from tf_agents.utils import common
from tf_agents.policies import policy_saver # add
from tf_agents.policies import py_tf_eager_policy # add
# add
import io
import os
import shutil
import tempfile
import zipfile
files = None
# add

file_path = curr_path + "/outputs/"
curr_time = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")  # 获取当前时间
# curr_path = os.path.dirname(os.path.abspath(__file__))  # 当前文件所在绝对路径
# tempdir = curr_path + "/C51_outputs/" + \
#             '/' + curr_time + '/models/'  # 保存模型的路径
# tb = tensorboardX.SummaryWriter()
tb = None
# tensorboard_callback = tensorboardX.(log_dir=log_dir, histogram_freq=1)
algo_name = "C51"  # 算法名称
env_name = 'C51_RobotOptEnv'  # 环境名称



class drl_optimization:
    def __init__(self):
        # self.test = 0
        self.robot = modular_robot_6dof()
        self.env = RobotOptEnv()
        # self.config = Config()

    def env_agent_config(self, cfg, algorithm, seed=1):
        ''' 创建环境和智能体
        '''
        # num_iterations = 15000 # @param {type:"integer"}

        # initial_collect_steps = 1000  # @param {type:"integer"}
        # collect_steps_per_iteration = 1  # @param {type:"integer"}
        # replay_buffer_capacity = 100000  # @param {type:"integer"}

        fc_layer_params = (100,)

        # batch_size = 64  # @param {type:"integer"}
        learning_rate = 1e-3  # @param {type:"number"}
        gamma = 0.99
        # log_interval = 200  # @param {type:"integer"}

        num_atoms = 51  # @param {type:"integer"}
        min_q_value = -205  # @param {type:"integer"}
        max_q_value = 110  # @param {type:"integer"}
        n_step_update = 2  # @param {type:"integer"}

        # num_eval_episodes = 10  # @param {type:"integer"}
        # eval_interval = 1000  # @param {type:"integer"}

        train_py_env = suite_gym.wrap_env(self.env)
        # eval_py_env = suite_gym.load(self.env)

        env = tf_py_environment.TFPyEnvironment(train_py_env)
        # eval_env = tf_py_environment.TFPyEnvironment(eval_py_env)

        categorical_q_net = categorical_q_network.CategoricalQNetwork(
            env.observation_spec(),
            env.action_spec(),
            num_atoms=num_atoms,
            fc_layer_params=fc_layer_params)

        dqn_network = QNetwork(
            env.observation_spec(),
            env.action_spec(),
            fc_layer_params=fc_layer_params)

        ddqn_network = QNetwork(
            env.observation_spec(),
            env.action_spec(),
            fc_layer_params=fc_layer_params)

        optimizer = tf.compat.v1.train.AdamOptimizer(learning_rate=learning_rate)
        self.global_step = tf.compat.v1.train.get_or_create_global_step()

        # train_step_counter = tf.Variable(0)
        # agent = DDQNAgent(self.config)  # 创建智能体
        if algorithm == 'DQN':
            agent = DqnAgent(
                env.time_step_spec(),
                env.action_spec(),
                q_network = dqn_network,
                optimizer = optimizer,
                n_step_update=n_step_update,
                td_errors_loss_fn = common.element_wise_squared_loss,
                gamma=gamma,
                train_step_counter = self.global_step)
            agent.initialize()
            rospy.loginfo("DRL algorithm init: %s", algorithm)
        elif algorithm == 'DDQN':
            agent = DdqnAgent(
                env.time_step_spec(),
                env.action_spec(),
                q_network = ddqn_network,
                optimizer = optimizer,
                n_step_update=n_step_update,
                td_errors_loss_fn = common.element_wise_squared_loss,
                gamma=gamma,
                train_step_counter = self.global_step)
            agent.initialize()
            rospy.loginfo("DRL algorithm init: %s", algorithm)
        elif algorithm == 'C51':
            agent = categorical_dqn_agent.CategoricalDqnAgent(
                env.time_step_spec(),
                env.action_spec(),
                categorical_q_network=categorical_q_net,
                optimizer=optimizer,
                min_q_value=min_q_value,
                max_q_value=max_q_value,
                n_step_update=n_step_update,
                td_errors_loss_fn=common.element_wise_squared_loss,
                gamma=gamma,
                train_step_counter=self.global_step)
            agent.initialize()
            rospy.loginfo("DRL algorithm init: %s", algorithm)

        # dqn_agent = DqnAgent(
        #     env.time_step_spec(),
        #     env.action_spec(),
        #     q_network = dqn_network,
        #     optimizer = optimizer,
        #     td_errors_loss_fn = common.element_wise_squared_loss,
        #     train_step_counter = self.global_step)

        # ddqn_agent = DdqnAgent(
        #     env.time_step_spec(),
        #     env.action_spec(),
        #     q_network = ddqn_network,
        #     optimizer = optimizer,
        #     td_errors_loss_fn = common.element_wise_squared_loss,
        #     train_step_counter = self.global_step)
        return env, agent


# class PlotConfig:
#     ''' 绘图相关参数设置
#     '''

#     def __init__(self) -> None:
#         self.algo_name = algo_name  # 算法名称
#         self.env_name = env_name  # 环境名称
#         self.device = torch.device(
#             "cuda" if torch.cuda.is_available() else "cpu")  # 检测GPU
#         self.result_path = curr_path + "/outputs/" + self.env_name + \
#             '/' + curr_time + '/results/'  # 保存结果的路径
#         self.model_path = curr_path + "/outputs/" + self.env_name + \
#             '/' + curr_time + '/models/'  # 保存模型的路径
#         self.save = True  # 是否保存图片

class RosTopic:
    def __init__(self):
        self.sub_taskcmd = rospy.Subscriber("/cal_command", cal_cmd, self.cmd_callback)
        self.sub_test_model_name = rospy.Subscriber("/tested_model_name", tested_model_name, self.test_model_name_callback)
        # Subscriber select dof and structure
        self.sub_arm_structure = rospy.Subscriber("/arn_structure", arm_structure, self.arm_structure_callback)
        self.cmd_run = 0
        self.arm_structure_dof = None
        self.DRL_algorithm = None
    def cmd_callback(self, data):
        self.cmd = data.cmd
        rospy.loginfo("I heard command is %s", data.cmd)
        if data.cmd == 22:
            rospy.sleep(10)
            self.cmd_run = 1
        if data.cmd == 23:
            rospy.sleep(10)
            self.cmd_run = 2
    def test_model_name_callback(self, data):
        self.test_model_name  = data.tested_model_name
        print(self.test_model_name)

    def arm_structure_callback(self, data):
        self.arm_structure_dof = data.DoF
        self.DRL_algorithm = data.structure_name
        # rospy.loginfo("DoF: %s", self.arm_structure_dof)
        # rospy.loginfo("DRL algorithm: %s", self.DRL_algorithm)

class Trainer:
    def __init__(self, agent, env, model_path):
        self.agent = agent
        self.env = env
        self.model_path = model_path
        # TODO: fixed
        # self.num_iterations = 15000 # @param {type:"integer"}
        # test
        self.num_iterations = 40000 # @param {type:"integer"}
        # self.initial_collect_steps = 1000  # @param {type:"integer"}
        # test
        self.initial_collect_steps = 1000  # @param {type:"integer"}
        self.collect_steps_per_iteration = 1  # @param {type:"integer"}
        self.replay_buffer_capacity = 100000  # @param {type:"integer"}

        self.random_policy = random_tf_policy.RandomTFPolicy(self.env.time_step_spec(),
                                                self.env.action_spec())

        self.replay_buffer = tf_uniform_replay_buffer.TFUniformReplayBuffer(
            data_spec=self.agent.collect_data_spec,
            batch_size=self.env.batch_size,
            max_length=self.replay_buffer_capacity)
        # add -----
        self.collect_driver = dynamic_step_driver.DynamicStepDriver(
            self.env,
            self.agent.collect_policy,
            observers=[self.replay_buffer.add_batch],
            num_steps=self.collect_steps_per_iteration)
        # Initial data collection
        self.collect_driver.run()

        self.checkpoint_dir = os.path.join(self.model_path, 'checkpoint')
        self.train_checkpointer = common.Checkpointer(
            ckpt_dir=self.checkpoint_dir,
            max_to_keep=1,
            agent=self.agent,
            policy=self.agent.policy,
            replay_buffer=self.replay_buffer,
            global_step=self.agent.train_step_counter
        )

        self.policy_dir = os.path.join(self.model_path, 'policy')
        self.tf_policy_saver = policy_saver.PolicySaver(agent.policy)
        # add -----
        self.batch_size = 64  # @param {type:"integer"}
        self.n_step_update = 2  # @param {type:"integer"}
        self.num_eval_episodes = 10  # @param {type:"integer"}
        self.log_interval = 200  # @param {type:"integer"}
        self.eval_interval = 1000  # @param {type:"integer"}

    def compute_avg_return(self, environment, policy, num_episodes=10):

        total_return = 0.0
        for _ in range(num_episodes):

            time_step = environment.reset()
            episode_return = 0.0

            while not time_step.is_last():
                action_step = policy.action(time_step)
                time_step = environment.step(action_step.action)
                episode_return += time_step.reward
            total_return += episode_return

        avg_return = total_return / num_episodes
        return avg_return.numpy()[0]

    def collect_step(self, environment, policy):
        time_step = environment.current_time_step()
        action_step = policy.action(time_step)
        next_time_step = environment.step(action_step.action)
        traj = trajectory.from_transition(time_step, action_step, next_time_step)

        # Add trajectory to the replay buffer
        self.replay_buffer.add_batch(traj)

        return next_time_step

    def train(self, pre_fr=0, train_eps = 300):
        rospy.loginfo("-------------數據採集------------")
        for _ in range(self.initial_collect_steps):
            time_step_collect = self.collect_step(self.env, self.random_policy)
            rospy.loginfo("initial_collect_steps: %s", _)
            # This loop is so common in RL, that we provide standard implementations of
            # these. For more details see the drivers module.

            # Dataset generates trajectories with shape [BxTx...] where
            # T = n_step_update + 1.
        dataset = self.replay_buffer.as_dataset(
            num_parallel_calls=3, sample_batch_size=self.batch_size,
            num_steps=self.n_step_update + 1).prefetch(3)

        iterator = iter(dataset)
        # (Optional) Optimize by wrapping some of the code in a graph using TF function.
        self.agent.train = common.function(self.agent.train)

        # Reset the train step
        self.agent.train_step_counter.assign(0)
        rospy.loginfo("Evaluate the agent's policy once before training.")
        # Evaluate the agent's policy once before training.
        avg_return = self.compute_avg_return(self.env, self.agent.policy, self.num_eval_episodes)
        returns = [avg_return]

        rospy.loginfo("-------------Train Start------------")
        for _ in range(self.num_iterations):

        # Collect a few steps using collect_policy and save to the replay buffer.
            for _ in range(self.collect_steps_per_iteration):
                time_step = self.collect_step(self.env, self.agent.collect_policy)

            # Sample a batch of data from the buffer and update the agent's network.
            experience, unused_info = next(iterator)
            train_loss = self.agent.train(experience)
            tb.add_scalar("/trained-model/Loss_per_frame/", float(train_loss.loss), _)
            step = self.agent.train_step_counter.numpy()

            # 開始tensorboard紀錄
            step_reward = time_step.reward.numpy()[0]
            tb.add_scalar("/trained-model/train_step_reward/", step_reward, step)
            # tb.add_scalar("/trained-model/Average_Return/", avg_return, step)

            if step % self.log_interval == 0:
                print('step = {0}: loss = {1}'.format(step, train_loss.loss))
                tb.add_scalar("/trained-model/loss_log/",  float(train_loss.loss), step)
            if step % self.eval_interval == 0:
                avg_return = self.compute_avg_return(self.env, self.agent.policy, self.num_eval_episodes)
                print('step = {0}: Average Return = {1:.2f}'.format(step, avg_return))
                tb.add_scalar("/trained-model/Average_Return/", avg_return, step)
                returns.append(avg_return)
        self.train_checkpointer.save(self.agent.train_step_counter)
        self.train_checkpointer.initialize_or_restore()
        self.agent.global_step = tf.compat.v1.train.get_global_step()

        self.tf_policy_saver.save(self.policy_dir)


class Tester(object):

    def __init__(self, env, model_path, drl_env_class, num_episodes=50, max_ep_steps=400, test_ep_steps=100):
        self.num_episodes = num_episodes
        self.max_ep_steps = max_ep_steps
        self.test_ep_steps = test_ep_steps
        # self.agent = agent
        self.model_path = model_path
        self.env = env
        self._best_episode_reward = 50
        self.drl_env_class = drl_env_class
        # self.agent.is_training = False
        # self.agent.load_weights(model_path)
        # self.policy = lambda x: agent.act(x)
        self.xlsx_outpath = "./xlsx/"

    def test(self, debug=True, visualize=True): # debug = true
        print("load model ")
        self.policy_dir = os.path.join(self.model_path, 'policy')
        saved_policy = tf.saved_model.load(self.policy_dir)

        excel_file = Workbook()
        sheet = excel_file.active
        i = 0
        # avg_reward = 0
        step = 0
        episode = 0
        avg_reward = 0
        state_traj = []
        num = 0
        for _ in range(self.num_episodes):
            time_step = self.env.reset()
            while not time_step.is_last():
                # step = self.agent.train_step_counter.numpy()
                action_step = saved_policy.action(time_step)
                time_step = self.env.step(action_step.action)
                step_reward = time_step.reward.numpy()[0]
                state = time_step.observation
                state_traj.append(state.numpy()[0])
                step += 1
                tb.add_scalar("/tested-model/test_step_reward/", step_reward, step)
            episode += 1
            episode_reward = step_reward
            
            for j in range(len(state.numpy()[0])):
                sheet.cell(row=i+1, column=j+1).value = state.numpy()[0][j] # state record 
            # self.motor_rated[1], self.motor_rated[2]] record
            sheet.cell(row=i+1, column=len(state.numpy()[0])+4).value = self.drl_env_class.motor_type_axis_2
            sheet.cell(row=i+1, column=len(state.numpy()[0])+5).value = self.drl_env_class.motor_type_axis_3
            # reward record
            sheet.cell(row=i + 1, column=len(state.numpy()[0])+2).value = episode_reward
            i = i + 1

            tb.add_scalar("/tested-model/test_episode_reward/", episode_reward, episode)
            avg_reward += episode_reward
            # TODO: 若獎勵大於50, 則儲存當前回合獎勵軌跡 
            if episode_reward >= self._best_episode_reward: 
                # TODO: state_traj
                excel_file_traj = Workbook()
                sheet_traj = excel_file_traj.active
                # 迭代矩陣的每一個元素，並填入工作表中
                for k in range(len(state_traj)):
                    for l in range(len(state_traj[k])):
                        sheet_traj.cell(row=k+1, column=l+1).value = state_traj[k][l]
                num = num + 1  # 要更改的数字
                new_str = "_{}_".format(num)
                file_name_traj = self.model_path + "/tested_state_traj" + new_str + curr_time +".xlsx"
                excel_file_traj.save(file_name_traj)
            # TODO: 清空
            else:
                state_traj = []

        avg_reward /= self.num_episodes
        # print("avg reward: %5f" % (avg_reward))
        rospy.loginfo('avg reward: {}'.format(avg_reward))

        # model_path = curr_path + '/train_results' + '/C51_outputs/' + str(arm_structure_dof) + \
        #         '/' + str(ros_topic.test_model_name) + '/models/'  # 保存模型的路径
        
        # add curr_time
        file_name = self.model_path + "/tested_reward_state_" + curr_time +".xlsx"
        excel_file.save(file_name)

        
        # TODO: save robot model structure, DH, urdf, stl

if __name__ == "__main__":
    rospy.init_node("optimization")
    a = 0
    # cfg = DQNConfig()
    cfg = 0
    arm_structure_dof = 0
    model_path = None
    drl = drl_optimization()
    # plot_cfg = PlotConfig()
    ros_topic = RosTopic()
    ddqn_train_eps = 1000  # 训练的回合数
    ddqn_test_eps = 100  # 测试的回合数
    
    
    with open(curr_path+'/drl_param.yaml', 'r') as f:
        config = yaml.load(f, Loader=yaml.Loader)
    op_function_flag = config['op_function']
    rospy.loginfo("Input op_function_flag: %s" % op_function_flag)    
    ros_topic.arm_structure_dof = config['arm_structure_dof']
    rospy.loginfo("Input arm_structure_dof: %d" % ros_topic.arm_structure_dof)
    ros_topic.DRL_algorithm = config['DRL_algorithm']
    rospy.loginfo("Input DRL_algorithm: %s" % ros_topic.DRL_algorithm)
    ros_topic.cmd_run = config['cmd_run']
    rospy.loginfo("Input cmd_run: %d" % ros_topic.cmd_run)
    ros_topic.test_model_name = config['test_model_name']
    rospy.loginfo("Input test_model_name: %s" % ros_topic.test_model_name)

    
    # train = Trainer()
    while not rospy.is_shutdown():
        # test all
        if op_function_flag == "case1":
            from RobotOptEnv_dynamixel import RobotOptEnv, RobotOptEnv_3dof, RobotOptEnv_5dof
        elif op_function_flag == "case2":
            from RobotOptEnv_dynamixel_v2 import RobotOptEnv, RobotOptEnv_3dof, RobotOptEnv_5dof
        elif op_function_flag == "case3":
            from RobotOptEnv_dynamixel_v3_motion import RobotOptEnv
        if ros_topic.arm_structure_dof == 6:
            drl.env = RobotOptEnv()
            rospy.loginfo('arm_structure_dof: {}'.format(ros_topic.arm_structure_dof))
            arm_structure_dof = ros_topic.arm_structure_dof
            ros_topic.arm_structure_dof = 0
        elif ros_topic.arm_structure_dof == 3:
            drl.env = RobotOptEnv_3dof()
            rospy.loginfo('arm_structure_dof: {}'.format(ros_topic.arm_structure_dof))
            arm_structure_dof = ros_topic.arm_structure_dof
            ros_topic.arm_structure_dof = 0
        elif ros_topic.arm_structure_dof == 5:
            drl.env = RobotOptEnv_5dof()
            rospy.loginfo('arm_structure_dof: {}'.format(ros_topic.arm_structure_dof))
            arm_structure_dof = ros_topic.arm_structure_dof
            ros_topic.arm_structure_dof = 0


        drl.env.op_dof = config['op_dof']
        rospy.loginfo("Input op_dof: %d" % drl.env.op_dof)
        
        drl.env.op_payload = config['op_payload']
        # rospy.loginfo("Input op_payload: %d" % drl.env.op_payload)
        rospy.loginfo('Input op_payload_position: {}'.format(drl.env.op_payload))
        
        drl.env.op_payload_position = config['op_payload_position']
        # rospy.loginfo("Input op_payload_position: %d" % drl.env.op_payload_position)
        rospy.loginfo('Input op_payload_position: {}'.format(drl.env.op_payload_position))

        drl.env.mission_time = config['mission_time']
        # rospy.loginfo("Input op_vel: %d" % drl.env.op_vel)
        rospy.loginfo('Input mission_time: {}'.format(drl.env.mission_time))
        
        drl.env.op_vel = config['op_vel']
        # rospy.loginfo("Input op_vel: %d" % drl.env.op_vel)
        rospy.loginfo('Input op_vel: {}'.format(drl.env.op_vel))
        
        drl.env.op_acc = config['op_acc']
        # rospy.loginfo("Input op_acc: %d" % drl.env.op_acc)
        rospy.loginfo('Input op_acc: {}'.format(drl.env.op_acc))
        
        drl.env.op_radius = config['op_radius']
        rospy.loginfo("Input op_radius: %d" % drl.env.op_radius)
        
        drl.env.op_weight = config['op_weight']
        rospy.loginfo("Input op_weight: %d" % drl.env.op_weight)
        
        drl.env.op_cost = config['op_cost']
        rospy.loginfo("Input op_cost: %d" % drl.env.op_cost)

        # 要開始時, 按下隨意鍵
        input_text = input("Enter some next: ")
        rospy.loginfo("Input text: %s" % input_text)
        
        tb = tensorboardX.SummaryWriter()


        if ros_topic.cmd_run == 1:
            ros_topic.cmd_run = 0
            if ros_topic.DRL_algorithm == 'DQN':
                model_path = curr_path + '/train_results' + '/DQN_outputs/' + op_function_flag + '/' +str(arm_structure_dof) + \
                '/' + curr_time + '/models/'  # 保存模型的路径
            elif ros_topic.DRL_algorithm == 'DDQN':
                model_path = curr_path + '/train_results' + '/DDQN_outputs/' + op_function_flag + '/' + str(arm_structure_dof) + \
                '/' + curr_time + '/models/'  # 保存模型的路径
            elif ros_topic.DRL_algorithm == 'C51':
                model_path = curr_path + '/train_results' + '/C51_outputs/' + op_function_flag + '/' + str(arm_structure_dof) + \
                '/' + curr_time + '/models/'  # 保存模型的路径
            
            # 訓練
            drl.env.model_select = "train"
            drl.env.point_Workspace_cal_Monte_Carlo()
            train_env, train_agent = drl.env_agent_config(cfg, ros_topic.DRL_algorithm, seed=1)
            # model_path = None
            train = Trainer(train_agent, train_env, model_path)
            train.train(train_eps = ddqn_train_eps)
            # # 測試
            drl.env.model_select = "test"
            save_result_path = curr_path + '/test_results' + '/C51_outputs/' + op_function_flag + '/' + str(arm_structure_dof) + \
                '/' + curr_time   # 保存結果的路径
            # plot_cfg.model_path = plot_cfg.model_path +'model_last.pkl'
            test_env, test_agent = drl.env_agent_config(cfg, ros_topic.DRL_algorithm, seed=10)
            test = Tester(test_env, model_path, drl.env, num_episodes = 200) # 20230309  change 300-> 200
            test.test()
            break

        # test tested_model
        if ros_topic.cmd_run == 2:
            ros_topic.cmd_run = 0
            # 測試
            # if op_function_flag == "case1":
            #     from RobotOptEnv_dynamixel import RobotOptEnv, RobotOptEnv_3dof, RobotOptEnv_5dof
            # elif op_function_flag == "case2":
            #     from RobotOptEnv_dynamixel_v2 import RobotOptEnv, RobotOptEnv_3dof, RobotOptEnv_5dof
            # elif op_function_flag == "case3":
            #     from RobotOptEnv_dynamixel_v3_motion import RobotOptEnv
            if ros_topic.DRL_algorithm == 'DQN':
                select_path = curr_path + '/train_results' + '/DQN_outputs/' + str(arm_structure_dof) + \
                '/' + str(ros_topic.test_model_name) + '/models/'  # 選擇模型的路径
            elif ros_topic.DRL_algorithm == 'DDQN':
                select_path = curr_path + '/train_results' + '/DDQN_outputs/' + str(arm_structure_dof) + \
                '/' + str(ros_topic.test_model_name) + '/models/'  # 選擇模型的路径
            elif ros_topic.DRL_algorithm == 'C51':
                select_path = curr_path + '/train_results' + '/C51_outputs/' + str(arm_structure_dof) + \
                '/' + str(ros_topic.test_model_name) + '/models/'  # 選擇模型的路径
            # TODO: fixed
            '''
            if ros_topic.DRL_algorithm == 'DQN':
                select_path = curr_path + '/train_results' + '/DQN_outputs/' + op_function_flag + str(arm_structure_dof) + \
                '/' + str(ros_topic.test_model_name) + '/models/'  # 選擇模型的路径
            elif ros_topic.DRL_algorithm == 'DDQN':
                select_path = curr_path + '/train_results' + '/DDQN_outputs/' + op_function_flag + str(arm_structure_dof) + \
                '/' + str(ros_topic.test_model_name) + '/models/'  # 選擇模型的路径
            elif ros_topic.DRL_algorithm == 'C51':
                select_path = curr_path + '/train_results' + '/C51_outputs/' + op_function_flag + str(arm_structure_dof) + \
                '/' + str(ros_topic.test_model_name) + '/models/'  # 選擇模型的路径
            '''
            drl.env.model_select = "test"
            # model_path = curr_path + '/test_results' + '/C51_outputs/' + str(arm_structure_dof) + \
            #     '/' + curr_time   # 保存結果的路径
            # model_path = select_path + str(ros_topic.test_model_name) +'/models/model_last.pkl'# test 20230102
            # model_path = select_path + str(ros_topic.test_model_name) +'/models/model_last.pkl'
            model_path = select_path
            test_env, test_agent = drl.env_agent_config(cfg, ros_topic.DRL_algorithm, seed=10)
            test = Tester(test_env, model_path, drl.env, num_episodes = 200) # 20230309  change 300-> 200 #20230326 mini-test 20
            test.test()
            break
        else:
            pass
